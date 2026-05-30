import requests
from bs4 import BeautifulSoup
from pydub import AudioSegment
import tempfile
import os
from openpyxl import Workbook
import re


def extract_timecoded_segments(soup: BeautifulSoup):

    table = soup.find("table", class_=["translation_table", "transcribe_form", "campl-table"])
    if not table:
        return None

    rows = table.select("tbody tr")
    segments = []
    has_any_timestamps = False

    for row in rows:

        time_el = row.find(class_="time-code")
        start_time = ""
        if time_el:
            start_time = (time_el.get("value") or time_el.get_text(strip=True) or "").strip()
            if start_time:
                has_any_timestamps = True


        tds = row.find_all("td", recursive=False)
        if len(tds) < 2:
            continue

        aramaic_td = tds[1]



        row_pieces = []
        seen_in_this_row = set()


        for el in aramaic_td.find_all(["div", "textarea"], class_="aramaic"):
            txt = (el.get_text(" ", strip=True) or el.get("value", "")).strip()
            if txt and txt not in seen_in_this_row:
                row_pieces.append(txt)
                seen_in_this_row.add(txt)


        if not row_pieces:
            raw_text = aramaic_td.get_text(" ", strip=True)
            if raw_text:
                row_pieces.append(raw_text)

        full_row_text = " ".join(row_pieces).strip()

        if full_row_text or start_time:
            segments.append({
                "start": start_time,
                "text": full_row_text
            })

    if not segments:
        return None


    has_real_text = any(s["text"].strip() for s in segments)


    if has_any_timestamps and not has_real_text:
        return None


    if has_any_timestamps:

        result = []
        for i in range(len(segments)):
            curr_start = segments[i]["start"]
            curr_end = segments[i + 1]["start"] if i + 1 < len(segments) else ""
            result.append((f"{curr_start}-{curr_end}", segments[i]["text"]))
        return result
    else:




        all_text_lines = [s["text"] for s in segments if s["text"]]

        full_text = "\n".join(all_text_lines)
        return [("", full_text)]


def extract_title_and_dialect(soup: BeautifulSoup) -> tuple[str | None, str | None]:
    h1 = soup.find("h1")
    if not h1:
        return None, None

    text = h1.get_text(strip=True)

    if ":" in text:
        dialect, title = text.split(":", 1)
        return title.strip(), dialect.strip()
    else:
        return text.strip(), None



ffmpeg_bin = r"C:\Users\alesi\ffmpeg\ffmpeg-8.0.1-essentials_build\bin"
os.environ["PATH"] += os.pathsep + ffmpeg_bin
AudioSegment.converter = os.path.join(ffmpeg_bin, "ffmpeg.exe")
AudioSegment.ffprobe = os.path.join(ffmpeg_bin, "ffprobe.exe")


base_folder = os.path.dirname(os.path.abspath(__file__))
texts_folder = os.path.join(base_folder, "data", "texts")
audios_folder = os.path.join(base_folder, "data", "audios")
unlabeled_audios_folder = os.path.join(base_folder, "unlabeled_data", "audios")

os.makedirs(unlabeled_audios_folder, exist_ok=True)
os.makedirs(texts_folder, exist_ok=True)
os.makedirs(audios_folder, exist_ok=True)


excel_path = os.path.join(base_folder, "metadata.xlsx")
wb = Workbook()
ws = wb.active
ws.title = "metadata"


no_tr_excel_path = os.path.join(base_folder, "no_transcript_metadata.xlsx")
wb_no_tr = Workbook()
ws_no_tr = wb_no_tr.active
ws_no_tr.title = "metadata"

labeled_headers = [
    "title of the text",
    "text file",
    "audio file length (seconds)",
    "source",
    "dialect"
]

unlabeled_headers = [
    "title of the text",
    "audio file",
    "audio file length (seconds)",
    "source",
    "dialect"
]

ws.append(labeled_headers)
ws_no_tr.append(unlabeled_headers)


main_url = "https://nena.ames.cam.ac.uk/dialects/225/audio"
resp_main = requests.get(main_url)
resp_main.raise_for_status()
soup_main = BeautifulSoup(resp_main.content, "html.parser")

audio_page_links = sorted({
    "https://nena.ames.cam.ac.uk" + a["href"]
    for a in soup_main.find_all("a", href=True)
    if a["href"].startswith("/audio/")
})

file_counter = 1
unlabeled_counter = 1

for page_url in audio_page_links:
    print(f"\nОбработка страницы: {page_url}")

    resp_page = requests.get(page_url)
    resp_page.raise_for_status()
    soup = BeautifulSoup(resp_page.content, "html.parser")

    segments = extract_timecoded_segments(soup)
    title_text, dialect = extract_title_and_dialect(soup)

    audio_link = soup.find("a", href=lambda x: x and x.endswith(".mp3"))

    labeled_wav_path = os.path.join(
        audios_folder,
        f"audio{file_counter}_urmi_labeled.wav"
    )

    unlabeled_wav_path = os.path.join(
        unlabeled_audios_folder,
        f"audio{unlabeled_counter}_urmi_unlabeled.wav"
    )

    text_path = os.path.join(
        texts_folder,
        f"text{file_counter}_urmi_labeled.xlsx"
    )

    success_audio = False
    success_text = False
    audio_length = None


    if audio_link:
        mp3_url = audio_link["href"]
        if not mp3_url.startswith("http"):
            mp3_url = "https://nena.ames.cam.ac.uk" + mp3_url

        with tempfile.NamedTemporaryFile(delete=False, suffix=".mp3") as tmp:
            tmp.write(requests.get(mp3_url).content)
            tmp_path = tmp.name

        sound = AudioSegment.from_mp3(tmp_path)
        sound.export(labeled_wav_path, format="wav")
        os.remove(tmp_path)

        total_seconds = int(sound.duration_seconds)
        audio_length = f"{total_seconds//3600:02d}:{(total_seconds%3600)//60:02d}:{total_seconds%60:02d}"
        success_audio = True


    if segments:
        segment_wb = Workbook()
        segment_ws = segment_wb.active
        segment_ws.title = "transcript"

        segment_ws.append(["timecode", "text"])

        for timecode, text in segments:
            segment_ws.append([timecode, text])

        segment_wb.save(text_path)
        success_text = True


    if success_audio and success_text:
        ws.append([
            title_text,
            os.path.basename(text_path),
            audio_length,
            "The North-Eastern Neo-Aramaic Database Project",
            dialect
        ])
        file_counter += 1

    elif success_audio and not success_text:
        os.rename(labeled_wav_path, unlabeled_wav_path)

        ws_no_tr.append([
            title_text,
            os.path.basename(unlabeled_wav_path),
            audio_length,
            "The North-Eastern Neo-Aramaic Database Project",
            dialect
        ])
        unlabeled_counter += 1


wb.save(excel_path)
wb_no_tr.save(no_tr_excel_path)
